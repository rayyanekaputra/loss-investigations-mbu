import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import glob

# Containers to store cumulative data
hpp_summary_list = []
profit_by_category_list = []
supplier_summary_list = []
ppn_summary_list = []

def generate_report(input_file):
    try:
        df = pd.read_excel(input_file)
    except Exception as e:
        print(f"Error reading {input_file}: {e}")
        return
    
    wb = Workbook()
    del wb['Sheet']

    def add_sheet_with_data(wb, sheet_name, df):
        ws = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    # Sheet 1: HPP vs Harga Beli Summary
    try:
        hpp_summary = df[['Nama Barang', '@Harga', '@Harga Beli']].copy()
        hpp_summary['Selisih HPP vs Harga Beli'] = hpp_summary['@Harga'] - hpp_summary['@Harga Beli']
        hpp_summary['% Selisih'] = (hpp_summary['Selisih HPP vs Harga Beli'] / hpp_summary['@Harga Beli']) * 100
        hpp_summary = hpp_summary.groupby('Nama Barang').mean().reset_index().round(2)
        hpp_summary['Source File'] = os.path.basename(input_file)
        hpp_summary_list.append(hpp_summary)
        add_sheet_with_data(wb, "HPP vs Harga Beli", hpp_summary)
    except KeyError as e:
        print(f"Missing columns in {input_file} for HPP summary: {e}")
    
    # Sheet 2: Detail HPP Analysis
    try:
        hpp_detail = df[['Nama Barang', 'Tanggal', 'Kuantitas', '@Harga', 'Total Harga', 
                         '@Harga Beli', 'Kuantitas Beli', 'Penjualan', 'Laba']].copy()
        hpp_detail['HPP'] = hpp_detail['Kuantitas'] * hpp_detail['@Harga']
        hpp_detail['HPP Based on Purchase'] = hpp_detail['Kuantitas'] * hpp_detail['@Harga Beli']
        hpp_detail['Selisih HPP'] = hpp_detail['HPP'] - hpp_detail['HPP Based on Purchase']
        hpp_detail['Laba Recalculated'] = hpp_detail['Penjualan'] - hpp_detail['HPP Based on Purchase']
        hpp_detail['Selisih Laba'] = hpp_detail['Laba Recalculated'] - hpp_detail['Laba']
        hpp_detail = hpp_detail.round(2)
        add_sheet_with_data(wb, "Detail HPP Analysis", hpp_detail)
    except KeyError as e:
        print(f"Missing columns in {input_file} for detailed HPP analysis: {e}")
    
    # Sheet 3: Profit by Category
    try:
        profit_by_category = df[['Nama Kategori Barang Barang & Jasa', 'Penjualan', 'Laba']].copy()
        profit_by_category['HPP'] = profit_by_category['Penjualan'] - profit_by_category['Laba']
        profit_by_category['Margin %'] = (profit_by_category['Laba'] / profit_by_category['Penjualan']) * 100
        profit_by_category = profit_by_category.groupby('Nama Kategori Barang Barang & Jasa').sum().reset_index().round(2)
        profit_by_category['Source File'] = os.path.basename(input_file)
        profit_by_category_list.append(profit_by_category)
        add_sheet_with_data(wb, "Profit by Category", profit_by_category)
    except KeyError as e:
        print(f"Missing columns in {input_file} for profit by category: {e}")

    # Sheet 4: Supplier Analysis
    try:
        supplier_analysis = df[['Nama Pemasok Faktur Pembelian Beli', '@Harga Beli', 'Kuantitas Beli']].copy()
        supplier_analysis['Total Pembelian'] = supplier_analysis['@Harga Beli'] * supplier_analysis['Kuantitas Beli']
        supplier_analysis = supplier_analysis.groupby('Nama Pemasok Faktur Pembelian Beli').agg({
            '@Harga Beli': 'mean',
            'Kuantitas Beli': 'sum',
            'Total Pembelian': 'sum'
        }).reset_index().round(2)
        supplier_analysis.rename(columns={
            '@Harga Beli': 'Rata-rata Harga Beli',
            'Kuantitas Beli': 'Total Kuantitas Dibeli',
            'Total Pembelian': 'Total Nilai Pembelian'
        }, inplace=True)
        supplier_analysis['Source File'] = os.path.basename(input_file)
        supplier_summary_list.append(supplier_analysis)
        add_sheet_with_data(wb, "Supplier Analysis", supplier_analysis)
    except KeyError as e:
        print(f"Missing columns in {input_file} for supplier analysis: {e}")

    # Sheet 5: PPN Analysis
    try:
        if 'Kena PPN' in df.columns and 'Kena PPN Beli' in df.columns:
            ppn_analysis = df[['Nama Barang', 'Kena PPN', 'Kena PPN Beli', 'Total Harga']].copy()
            ppn_analysis['PPN Penjualan'] = ppn_analysis.apply(lambda x: x['Total Harga'] * 0.11 if x['Kena PPN'] == 'Ya' else 0, axis=1)
            ppn_analysis['PPN Pembelian'] = ppn_analysis.apply(lambda x: x['Total Harga'] * 0.11 if x['Kena PPN Beli'] == 'Ya' else 0, axis=1)
            ppn_analysis['PPN Terutang'] = ppn_analysis['PPN Penjualan'] - ppn_analysis['PPN Pembelian']
            ppn_summary = ppn_analysis.groupby('Nama Barang').sum().reset_index().round(2)
            ppn_summary['Source File'] = os.path.basename(input_file)
            ppn_summary_list.append(ppn_summary)
            add_sheet_with_data(wb, "PPN Analysis", ppn_summary)
    except KeyError as e:
        print(f"Missing columns in {input_file} for PPN analysis: {e}")

    output_file = os.path.splitext(os.path.basename(input_file))[0] + "_ANALYSIS_REPORT.xlsx"
    output_path = os.path.join(os.path.dirname(input_file), output_file)
    wb.save(output_path)
    print(f"Report generated successfully: {output_path}")

def generate_cumulative_summary():
    if not hpp_summary_list and not profit_by_category_list and not supplier_summary_list and not ppn_summary_list:
        print("No cumulative data to summarize.")
        return
    
    wb = Workbook()
    del wb['Sheet']

    def add_sheet(sheet_name, df):
        ws = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    if hpp_summary_list:
        hpp_all = pd.concat(hpp_summary_list, ignore_index=True)
        add_sheet("All HPP Summary", hpp_all)

    if profit_by_category_list:
        profit_all = pd.concat(profit_by_category_list, ignore_index=True)
        add_sheet("All Profit by Category", profit_all)

    if supplier_summary_list:
        supplier_all = pd.concat(supplier_summary_list, ignore_index=True)
        add_sheet("All Supplier Summary", supplier_all)

    if ppn_summary_list:
        ppn_all = pd.concat(ppn_summary_list, ignore_index=True)
        add_sheet("All PPN Summary", ppn_all)

    output_path = "./BAEKMI/ALL_MONTHS_SUMMARY.xlsx"
    wb.save(output_path)
    print(f"\nCumulative summary report saved at: {output_path}")

def process_all_files():
    input_folder = os.path.normpath("./BAEKMI/")
    os.makedirs(input_folder, exist_ok=True)
    input_files = glob.glob(os.path.join(input_folder, "*_merge_with_purchasing_*.xlsx"))
    
    if not input_files:
        print(f"No matching files found in {input_folder}")
        return
    
    for input_file in input_files:
        print(f"\nProcessing file: {input_file}")
        generate_report(input_file)
    
    generate_cumulative_summary()

if __name__ == "__main__":
    process_all_files()
    print("\nAll files processed!")
